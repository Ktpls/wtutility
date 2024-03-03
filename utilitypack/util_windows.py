from .util_solid import *
from .util_np import *
from .util_ocv import *
from ctypes import windll, byref, c_ubyte
from ctypes.wintypes import RECT, HWND
from zipfile import ZipFile
import win32api
import win32con
import win32gui
import win32ui
import subprocess
import regex

"""
windows
"""


class screenshoter:
    # reconstructed with no mfc

    def __init__(self, hwnd=0):
        self.wthwnd = hwnd
        # r = RECT()
        # windll.user32.GetClientRect(self.hwnd, byref(r))
        # self.res=[r.right,r.bottom]
        self.res = [1920, 1080]
        self.wtdc = windll.user32.GetDC(self.wthwnd)
        self.mydc = windll.gdi32.CreateCompatibleDC(self.wtdc)
        self.mybitmap = windll.gdi32.CreateCompatibleBitmap(
            self.wtdc, self.res[0], self.res[1]
        )
        windll.gdi32.SelectObject(self.mydc, self.mybitmap)

    def __del__(self):
        windll.gdi32.DeleteObject(self.mybitmap)
        windll.gdi32.DeleteObject(self.mydc)
        windll.user32.ReleaseDC(self.wthwnd, self.wtdc)

    def shot(self):
        if (
            windll.gdi32.BitBlt(
                self.mydc,
                0,
                0,
                self.res[0],
                self.res[1],
                self.wtdc,
                0,
                0,
                win32con.SRCCOPY,
            )
            == 0
        ):
            raise Exception("bad shot, {}".format(windll.kernel32.GetLastError()))
        # 截图是BGRA排列，因此总元素个数需要乘以4
        total_bytes = self.res[0] * self.res[1] * 4
        buffer = bytearray(total_bytes)
        byte_array = c_ubyte * total_bytes
        windll.gdi32.GetBitmapBits(
            self.mybitmap, total_bytes, byte_array.from_buffer(buffer)
        )
        return np.frombuffer(buffer, dtype=np.uint8).reshape(
            self.res[1], self.res[0], 4
        )

    def shotbgr(self):
        return self.shot()[:, :, :3]

    def shotgray(self):
        return cv.cvtColor(self.shot(), cv.COLOR_BGRA2GRAY)


def activeWindow(hwnd):  # 窗口置顶
    win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
    win32gui.SetWindowPos(
        hwnd,
        win32con.HWND_TOP,
        0,
        0,
        0,
        0,
        win32con.SWP_NOSIZE | win32con.SWP_SHOWWINDOW,
    )
    win32gui.SetForegroundWindow(hwnd)


def bootAsAdmin(file):
    # pass in __file__
    quotedpy = '"' + file + '"'
    ret = ctypes.windll.shell32.ShellExecuteW(
        None, "runas", sys.executable, quotedpy, None, 1
    )


def setadmin(file):
    def is_admin():
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    if not is_admin():
        bootAsAdmin(file)
        exit()


class fullScrHUD:
    """
    #init
    hud=fullScrHUD()
    hud.setup()
    ...
    #main loop
    #preparing to show
    hud.clear
    #init m1
    m1=hud.getblankscreen()
    #draw sth on m1
    #then add m1 to hud
    hud.addcontent(m1)
    #or overwrite directly
    hud.writecontent(m1)
    #do the same for m2...
    ...
    #ask hud to show
    hud.update()
    ...
    #on exit
    hud.stop()
    """

    def __init__(self) -> None:
        self.resolution = [1080, 1920]
        self.terminate = False
        self.hwnd = 0
        self.m2show = self.getblankscreenwithalfa()
        self.m2draw = self.getblankscreenwithalfa()
        # set m2draw

    @FunctionalWrapper
    def setup(self):
        def WndProc(hwnd, msg, wParam, lParam):
            if msg == win32con.WM_PAINT:
                rect = win32gui.GetClientRect(hwnd)
                hdc, ps = win32gui.BeginPaint(hwnd)
                # background
                # hbr=win32gui.CreateSolidBrush(0x0000000)
                # win32gui.FillRect(hdc,rect,hbr)
                # win32gui.DeleteObject(hbr)

                w = self.resolution[1]
                h = self.resolution[0]
                mfcDC = win32ui.CreateDCFromHandle(hdc)
                hcdc = mfcDC.CreateCompatibleDC()
                BitMap = win32ui.CreateBitmap()
                BitMap.CreateCompatibleBitmap(mfcDC, w, h)
                ctypes.WinDLL("gdi32.dll").SetBitmapBits(
                    BitMap.GetHandle(), w * h * 4, self.m2show.tobytes()
                )
                hcdc.SelectObject(BitMap)
                # myblendfunc=(win32con.AC_SRC_OVER,0,255,win32con.AC_SRC_ALPHA)
                # win32gui.AlphaBlend(hdc,0,0,w, h , hcdc.GetHandleAttrib(), 0,0,w, h,myblendfunc)
                win32gui.BitBlt(
                    hdc, 0, 0, w, h, hcdc.GetHandleAttrib(), 0, 0, win32con.SRCCOPY
                )
                win32gui.DeleteObject(BitMap.GetHandle())
                hcdc.DeleteDC()
                win32gui.EndPaint(hwnd, ps)
            elif msg == win32con.WM_SIZE:
                win32gui.InvalidateRect(hwnd, None, True)
            elif msg == win32con.WM_DESTROY:
                win32gui.PostQuitMessage(0)
            if self.terminate:
                win32gui.PostQuitMessage(0)
            return win32gui.DefWindowProc(hwnd, msg, wParam, lParam)

        def mainloop():
            wc = win32gui.WNDCLASS()
            hbrush = win32gui.CreateSolidBrush(0)
            wc.hbrBackground = hbrush
            wc.lpszClassName = "Python on Windows"
            wc.lpfnWndProc = WndProc
            reg = win32gui.RegisterClass(wc)

            hwnd = win32gui.CreateWindow(
                reg,
                "Python",
                win32con.WS_POPUP,
                0,
                0,
                self.resolution[1],
                self.resolution[0],
                0,
                0,
                0,
                None,
            )
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
            win32gui.UpdateWindow(hwnd)

            win32gui.SetWindowLong(
                hwnd,
                win32con.GWL_EXSTYLE,
                win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)
                + win32con.WS_EX_LAYERED
                + win32con.WS_EX_NOACTIVATE,
            )
            win32gui.SetLayeredWindowAttributes(hwnd, 0, 0, win32con.LWA_COLORKEY)
            win32gui.SetWindowPos(
                hwnd,
                win32con.HWND_TOPMOST,
                0,
                0,
                0,
                0,
                win32con.SWP_SHOWWINDOW + win32con.SWP_NOSIZE + win32con.SWP_NOMOVE,
            )
            windll.user32.SetWindowDisplayAffinity(
                hwnd, 0x11
            )  # WDA_EXCLUDEDFROMCAPUTURE
            # can not use WDA_MONITOR, or get totally black screen

            self.hwnd = hwnd
            win32gui.PumpMessages()

        t1 = threading.Thread(target=mainloop, args=())
        t1.start()
        time.sleep(1)

    @FunctionalWrapper
    def writecontent(self, lt, content):
        self.m2draw[
            lt[0] : lt[0] + content.shape[0],
            lt[1] : lt[1] + content.shape[1],
            : content.shape[2],
        ] = content

    @FunctionalWrapper
    def clear(self):
        self.m2draw[:, :, :] = 0

    def getblankscreen(self):
        return np.zeros(
            self.resolution
            + [
                3,
            ],
            np.uint8,
        )

    def getblankscreenwithalfa(self):
        return np.zeros(
            self.resolution
            + [
                4,
            ],
            np.uint8,
        )

    @FunctionalWrapper
    def addcontentwithalfa(self, m):
        self.m2draw += m

    @FunctionalWrapper
    def addcontent(self, m):
        alp = 255 * np.ones_like(m[:, :, 0:1])
        self.m2draw += np.concatenate((m, alp), 2)

    @FunctionalWrapper
    def update(self):
        # swap
        tmp = self.m2show
        self.m2show = self.m2draw
        self.m2draw = tmp

        # needless to clear cuz entire screen will be set
        win32gui.InvalidateRect(self.hwnd, None, False)

    @FunctionalWrapper
    def stop(self):
        self.terminate = True
        self.update()

    @staticmethod
    def reverseMat2FitWindowsCoor(m):
        return np.flip(m, axis=0)


def isKBDown(k):
    return bool(win32api.GetAsyncKeyState(k) and 0x1)


def isKBDownNow(k):
    return win32api.GetAsyncKeyState(k) and 0x8000


class TranslateHotKey:
    dct = {
        win32con.VK_CONTROL: "Ctrl",
        win32con.VK_SHIFT: "Shift",
        win32con.VK_MENU: "Alt",
        win32con.VK_RCONTROL: "RCtrl",
        win32con.VK_RSHIFT: "RShift",
        win32con.VK_RMENU: "RAlt",
        win32con.VK_LCONTROL: "LCtrl",
        win32con.VK_LSHIFT: "LShift",
        win32con.VK_LMENU: "LAlt",
        win32con.VK_BACK: "Backspace",
        win32con.VK_RETURN: "Enter",
        win32con.VK_ESCAPE: "Esc",
        win32con.VK_SPACE: "Space",
        win32con.VK_TAB: "Tab",
        win32con.VK_F1: "F1",
        win32con.VK_F2: "F2",
        win32con.VK_F3: "F3",
        win32con.VK_F4: "F4",
        win32con.VK_F5: "F5",
        win32con.VK_F6: "F6",
        win32con.VK_F7: "F7",
        win32con.VK_F8: "F8",
        win32con.VK_F9: "F9",
        win32con.VK_F10: "F10",
        win32con.VK_F11: "F11",
        win32con.VK_F12: "F12",
        win32con.VK_NUMPAD0: "Numpad0",
        win32con.VK_NUMPAD1: "Numpad1",
        win32con.VK_NUMPAD2: "Numpad2",
        win32con.VK_NUMPAD3: "Numpad3",
        win32con.VK_NUMPAD4: "Numpad4",
        win32con.VK_NUMPAD5: "Numpad5",
        win32con.VK_NUMPAD6: "Numpad6",
        win32con.VK_NUMPAD7: "Numpad7",
        win32con.VK_NUMPAD8: "Numpad8",
        win32con.VK_NUMPAD9: "Numpad9",
        win32con.VK_UP: "Up",
        win32con.VK_LEFT: "Left",
        win32con.VK_DOWN: "Down",
        win32con.VK_RIGHT: "Right",
        0xC0: "~",
    }

    @staticmethod
    def __call__(k):
        if (k >= ord("A") and k <= ord("Z")) or (k >= ord("0") and k <= ord("9")):
            return chr(k)
        assert k in TranslateHotKey.dct, "unknown key"
        return TranslateHotKey.dct[k]


class HotkeyManager:
    """
    to piorer ctrl+c than c
    responde no c after doing ctrl+c
    key is given by win32con.VK_*,
        for letters, use ord([letter's UPPER case])
        for numbers, use ord([number])
    issue:
    1. ctrl+a,ctrl+a and ctrl+a,ctrl+s
        fastly press ctrl+a, and ctrl+s, could cause slightly overlap whose key state is [ctrl,a,s] physically
        that is, [ctrl,a]->[ctrl,a,s]->[ctrl,s]
        so at the second state, both [ctrl+a,ctrl+a] and [ctrl+a,ctrl+s] are triggered
        not severe problem
    """

    @dataclasses.dataclass
    class ContiniousCallHandler:
        prevState: typing.Dict[int, bool] = dataclasses.field(default_factory=dict)
        countiousPressTime: int = 0
        startRepeatPeriod: int = 10
        useControlOnContiniousPress: bool = True

        def updateState(self, newState):
            if DictEq(self.prevState, newState):
                self.countiousPressTime += 1
                return self.countiousPressTime >= self.startRepeatPeriod
            else:
                self.countiousPressTime = 0
                self.prevState = newState
                return True

    class hotkeytask:
        key: list[list[int]]

        def __init__(
            self,
            key: int | typing.Iterable[int] | typing.Iterable[typing.Iterable[int]],
            foo: typing.Callable[[], None],
            continiousPress: bool = None,
        ) -> None:
            self.key = HotkeyManager.hotkeytask.formalize_key_param(key)
            # self.key is like [keyset1=[key1, key2], keyset2=[key3, key4]]
            self.foo = foo
            self.continiousPress = continiousPress if continiousPress else False
            self.stage = 0
            self.keyChangeOnStageChange = [
                (
                    ListEq(self.key[i], self.key[i - 1])
                    if i != 0
                    else ListEq(self.key[0], self.key[-1])
                )
                for i, k in enumerate(self.key)
            ]

        @staticmethod
        def formalize_key_param(key):
            if not isinstance(key, typing.Iterable):
                key = [[key]]
            elif not isinstance(key[0], typing.Iterable):
                key = [key]
            else:
                key = key
            return key

        @staticmethod
        def getKeyRepr(key):
            key = HotkeyManager.hotkeytask.formalize_key_param(key)
            if len(key) == 1:
                # only one stage
                return " + ".join([TranslateHotKey()(k) for k in key[0]])
            else:
                return "\n".join(
                    [
                        [
                            f"Stage {i}: {' + '.join([TranslateHotKey()(k) for k in ks])}"
                            for i, ks in enumerate(key)
                        ]
                    ]
                )

        def tryRespond(self, respond: bool, anyRespondingExceptThis: bool) -> bool:
            """
            ret: if key changed
            """
            if respond:
                if self.stage == len(self.key) - 1:
                    # progress completed
                    self.stage = 0
                    self.foo()
                else:
                    self.stage += 1
            elif anyRespondingExceptThis:
                self.stage = 0
            return self.keyChangeOnStageChange[self.stage]

        def GetNowKey(self):
            return self.key[self.stage]

    @dataclasses.dataclass
    class Key:
        code: int

        def GetKeyDown(self):
            return isKBDown(self.code)

    def __init__(self, hotkeytasklist: list[hotkeytask]):
        keyconcerned = [hka.key for hka in hotkeytasklist]
        keyconcerned = list(itertools.chain.from_iterable(keyconcerned))
        keyconcerned = list(itertools.chain.from_iterable(keyconcerned))
        keyconcerned = Deduplicate(keyconcerned)
        self.kc = [HotkeyManager.Key(k) for k in keyconcerned]

        # clear all previous state
        [k.GetKeyDown() for k in self.kc]

        self.hktl = hotkeytasklist
        self.cch = HotkeyManager.ContiniousCallHandler()
        self.__calcPriorInfo()

    def __calcPriorInfo(self):
        """
        costly!!!
        at m^2n^2, where m is #hotkeytask, n is #key of hotkeytask
        """

        def piorered(a: HotkeyManager.hotkeytask, b: HotkeyManager.hotkeytask):
            def include(a: HotkeyManager.hotkeytask, b: HotkeyManager.hotkeytask):
                for k in b.GetNowKey():
                    if k not in a.GetNowKey():
                        return False
                return True

            # a>b and b<a, not equal
            return include(a, b) and not include(b, a)

        self.piorinfo = [
            [
                aidx
                for aidx, a in enumerate(self.hktl)
                if aidx != bidx and piorered(a, b)
            ]
            for bidx, b in enumerate(self.hktl)
        ]

    def decideAllHotKey(self) -> list[bool]:
        keystate = {k.code: k.GetKeyDown() for k in self.kc}
        cchblocked = not self.cch.updateState(keystate)

        class respondstate(enum.Enum):
            false = 0
            true = 1
            unknown = 2

        respondtable = [
            (
                respondstate.unknown
                if not cchblocked or hk.continiousPress
                else respondstate.false
            )
            for hk in self.hktl
        ]

        def decideRespondState(i):
            # checked
            if respondtable[i] != respondstate.unknown:
                return

            # all key pressed
            if all([keystate[k] for k in self.hktl[i].GetNowKey()]):
                # didnt check piored, check it
                [
                    decideRespondState(p)
                    for p in self.piorinfo[i]
                    if respondtable[p] == respondstate.unknown
                ]

                # no piored responded
                if all(
                    [respondtable[p] == respondstate.false for p in self.piorinfo[i]]
                ):
                    respondtable[i] = respondstate.true
                else:
                    respondtable[i] = respondstate.false
            else:
                # not respond this
                respondtable[i] = respondstate.false

        for hkidx, hk in enumerate(self.hktl):
            decideRespondState(hkidx)

        assert all([rt != respondstate.unknown for rt in respondtable])
        r = [
            respondtable[hkidx] == respondstate.true
            for hkidx, hk in enumerate(self.hktl)
        ]
        return [
            respondtable[hkidx] == respondstate.true
            for hkidx, hk in enumerate(self.hktl)
        ]

    def doAllDecidedKey(self, decideresult, throwonerr=False, printonerr=False):
        AnyRespondingExceptThis = [
            any([(rb if i != j else False) for j, rb in enumerate(decideresult)])
            for i, ra in enumerate(decideresult)
        ]
        anyKeyChanged = False
        for i in range(len(decideresult)):
            try:
                thisKeyChanged = self.hktl[i].tryRespond(
                    decideresult[i], AnyRespondingExceptThis[i]
                )
                anyKeyChanged = anyKeyChanged or thisKeyChanged
            except Exception as e:
                if printonerr:
                    traceback.print_exc()
                if throwonerr:
                    raise e
        if anyKeyChanged:
            self.__calcPriorInfo()

    @dataclasses.dataclass
    class InputSession:
        @dataclasses.dataclass
        class SessionInstance:
            class SessionEndType(enum.Enum):
                UNSPECIFIED = 0
                OK = 1
                CANCEL = 2

            FooSessionDoneCallback: typing.Callable[
                [typing.Type["HotkeyManager.InputSession.SessionInstance"]], None
            ] = None
            content: str = ""
            sessionEndType: SessionEndType = SessionEndType.UNSPECIFIED

            def append(self, extraContent: str):
                self.content += extraContent

            def backSpace(self):
                self.content = self.content[:-1]

            def putup(self, bulletin: BulletinBoard):
                bulletin.putup(self.content)

        # foo that sets hkm and returns older hkm
        FooSwapHKM: typing.Callable[["HotkeyManager"], "HotkeyManager"]
        bulletin: BulletinBoard
        RunningSessionInstance: SessionInstance = dataclasses.field(
            default_factory=SessionInstance
        )
        hotkeymanagerStack: list["HotkeyManager"] = dataclasses.field(
            default_factory=list
        )

        class InputTypeEnabled(enum.Enum):
            NUMBER = 0
            LETTER = 1

        def __GetHotkeyReg(self, ite: list[InputTypeEnabled]):
            @dataclasses.dataclass
            class KeyMapping:
                char: str
                key: typing.Tuple[int]

            AllKeyMapping = [
                *(
                    [KeyMapping(k, (ord(k),)) for i, k in enumerate("0123456789")]
                    if HotkeyManager.InputSession.InputTypeEnabled.NUMBER in ite
                    else []
                ),
                *(
                    [
                        *[
                            KeyMapping(k, (ord(k.upper()),))
                            for i, k in enumerate("abcdefghijklmnopqrstuvwxyz")
                        ],
                        *[
                            KeyMapping(k, (win32con.VK_SHIFT, ord(k.upper())))
                            for i, k in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
                        ],
                    ]
                    if HotkeyManager.InputSession.InputTypeEnabled.LETTER in ite
                    else []
                ),
            ]
            KeyIndexed = {k.key: k.char for k in AllKeyMapping}

            def procKey(k):
                self.RunningSessionInstance.append(KeyIndexed.get(k, "?"))
                self.RunningSessionInstance.putup(self.bulletin)

            HotkeyReg = [
                HotkeyManager.hotkeytask(
                    key=[k.key], foo=functools.partial(procKey, k.key)
                )
                for k in AllKeyMapping
            ]

            def backSpace():
                self.RunningSessionInstance.backSpace()
                self.RunningSessionInstance.putup(self.bulletin)

            HotkeyReg.append(
                HotkeyManager.hotkeytask(
                    key=[win32con.VK_BACK],
                    foo=backSpace,
                )
            )

            def OutFromSession(
                endType: HotkeyManager.InputSession.SessionInstance.SessionEndType,
            ):
                self.RunningSessionInstance.sessionEndType = endType
                self.RunningSessionInstance.FooSessionDoneCallback(
                    self.RunningSessionInstance
                )
                old = self.hotkeymanagerStack.pop()
                inputer = self.FooSwapHKM(old)

            HotkeyReg.extend(
                [
                    HotkeyManager.hotkeytask(
                        key=[win32con.VK_ESCAPE],
                        foo=functools.partial(
                            OutFromSession,
                            HotkeyManager.InputSession.SessionInstance.SessionEndType.CANCEL,
                        ),
                    ),
                    HotkeyManager.hotkeytask(
                        key=[win32con.VK_RETURN],
                        foo=functools.partial(
                            OutFromSession,
                            HotkeyManager.InputSession.SessionInstance.SessionEndType.OK,
                        ),
                    ),
                ]
            )

            return HotkeyReg

        def IntoSession(
            self, callback, allowedInputType: list[InputTypeEnabled] = None
        ):
            self.RunningSessionInstance = HotkeyManager.InputSession.SessionInstance(
                callback
            )
            if not allowedInputType:
                allowedInputType = [
                    HotkeyManager.InputSession.InputTypeEnabled.NUMBER,
                    HotkeyManager.InputSession.InputTypeEnabled.LETTER,
                ]
            inputer = HotkeyManager(self.__GetHotkeyReg(allowedInputType))
            old = self.FooSwapHKM(inputer)
            self.hotkeymanagerStack.append(old)


def NormalizeCrlf(s: str):
    return s.replace("\r\n", "\n").replace("\r", "\n")


"""
xls
"""

import openpyxl as opx


def save_list_to_xls(
    data_list: list[list | tuple | typing.Any], filename, sheetname=None
):
    # Create a new workbook
    wb = opx.Workbook()

    # Select the active worksheet
    if sheetname is None:
        ws = wb.active
    else:
        ws = wb.create_sheet(sheetname)

    # Iterate over the list and write each item to a new row
    for row, rowcontent in enumerate(data_list):
        rowcontent=NormalizeIterableOrSingleArgToIterable(rowcontent)
        for col, item in enumerate(rowcontent):
            ws.cell(row=row + 1, column=col + 1, value=item)

    # Save the workbook to the specified filename
    wb.save(filename)


def Xls2ListList(path=None, sheetname=None, killNones=True):
    if path is None:
        path = r"eles.in.xlsx"
    xls = opx.load_workbook(path)
    if sheetname is None:
        sheet = xls.active
    else:
        sheet = xls[sheetname]

    ret = [[ele.value for ele in ln] for ln in (sheet.rows)]
    if killNones:
        ret = [l for l in ret if any([e is not None for e in l])]
    return ret


@dataclasses.dataclass
class Rhythm:
    @dataclasses.dataclass
    class BeepTone:
        freq: int
        dur: int = 100

    tones: list["Rhythm.BeepTone"]

    @staticmethod
    def fromString(s: str, default_dur: int = 100):
        lines = regex.split(r"\s+", NormalizeCrlf(s))
        tones = []
        for l in lines:
            arg = l.split(",")
            tones.append(
                Rhythm.BeepTone(
                    int(arg[0]), dur=int(arg[1]) if len(arg) > 1 else default_dur
                )
            )
        return Rhythm(tones)

    def play(self):
        for t in self.tones:
            win32api.Beep(t.freq, t.dur)

    @StoppableSomewhat.EasyUse(
        strategy_runonrunning=StoppableSomewhat.StrategyRunOnRunning.stop_and_rerun,
        implType=StoppableThread,
    )
    def asyncPlay(selfStoppable: StoppableSomewhat, selfRythm: "Rhythm"):
        selfRythm.play()


class Rhythms:
    Success = Rhythm.fromString("1000 500 1000", default_dur=100)
    Error = Rhythm.fromString("1000,1000 500,1000", default_dur=100)
    Cancel = Rhythm.fromString("1000 500 500", default_dur=100)
    Notify = Rhythm.fromString("500", default_dur=100)
    GoodNotify = Rhythm.fromString("1000", default_dur=100)
    BadNotify = Rhythm.fromString("500 500 500", default_dur=100)
    Reboot = Rhythm.fromString("500 750 400", default_dur=100)


def ReadFileInZip(zipf, filename: str | list[str] | tuple[str]):
    zipf = ZipFile(zipf)
    singleFile = not isinstance(filename, (tuple, list))
    if singleFile:
        filename = [filename]
    file = [zipf.read(f) for f in filename]
    if singleFile:
        return file[0]
    return file


class WifiRefresher:
    @staticmethod
    def __getNowWlanProfileName():
        output = subprocess.check_output(
            "netsh wlan show interface", shell=True
        ).decode("GBK")
        m = regex.findall(
            r"(?<=^\s*SSID\s*:\s)(.+?)(?=\s*$)", output, flags=regex.MULTILINE
        )
        assert len(m) == 1
        return m[0]

    def __init__(self):
        self.name = WifiRefresher.__getNowWlanProfileName()

    def setOn(self):
        subprocess.Popen(f'netsh wlan connect name="{self.name}"')
        subprocess.Popen(
            f'netsh wlan set profileparameter name="{self.name}" connectionMode=auto'
        )
        # set auto so it will be auto connected the next time u boot

    def setOff(self):
        subprocess.Popen(f"netsh wlan disconnect")
