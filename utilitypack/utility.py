from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
import enum
from time import sleep
from typing import Dict, List, Callable, Iterable, Any
import ctypes
import dataclasses
import itertools
import math
import os
import random
import re
import sys
import threading
import time
import traceback
import typing

"""
solid
"""


def deduplicate(l: typing.List):
    return list(set(l))


def digitsof(s: str):
    return "".join(list(filter(str.isdigit, list(s))))


def numinstr(s: str):
    # wont consider negative
    s = digitsof(s)
    return int(s) if len(s) > 0 else 0


class logger:
    def __init__(self, path):
        self.path = path
        # wont fail
        dir = os.path.dirname(path)
        if not os.path.exists(dir):
            os.makedirs(dir)
        self.f = open(path, "wb+")

    def log(self, content):
        self.f.write((content + "\n").encode("utf8"))
        # self.f.flush()

    def __del__(self):
        self.f.close()

    def __call__(self, content):
        self.log(content)


def quickSummonCard(inteprob):
    # faster summon using division
    pos = random.random() * inteprob[-1]
    section = [0, len(inteprob)]

    def compare(n):
        # compare pos with section[n]
        if pos > inteprob[n]:
            return 1
        else:  # pos<=section[n]
            if pos > (inteprob[n - 1] if n >= 1 else 0):
                return 0
            else:
                return -1

    while True:
        mid = int((section[1] + section[0]) * 0.5)
        compresult = compare(mid)
        if compresult == 1:
            section[0] = mid + 1
        elif compresult == -1:
            section[1] = mid
        else:  # compresult==0
            return mid


class toast:
    messagelist = []

    def sendmessage(self, content, peroid):
        self.messagelist.append(
            {"ctt": content, "st": time.perf_counter(), "per": peroid}
        )

    def updatemsglist(self):
        nowtime = time.perf_counter()

        def filterfoo(m):
            return m["per"] > m["ctt"] - nowtime

        messagelist = [m for m in messagelist if filterfoo(m)]
        msgs = [m["ctt"] for m in messagelist]

    def getallmsg(self):
        self.updatemsglist()
        return "\n".join(self.messagelist)


class bulletinBoard:
    # new msg covers the lasts
    def __init__(self, idlecontent):
        self.idlecontent = idlecontent
        self.content = ""
        self.overduetime = time.perf_counter()

    def putup(self, content, timeout=10):
        self.content = content
        self.overduetime = time.perf_counter() + timeout

    def read(self):
        if time.perf_counter() < self.overduetime:
            return self.content
        else:
            return self.idlecontent


def outputlines2mat(m, pos, content, lineheight=25, textcolor=[255, 255, 255]):
    m = m.copy()
    line = content.split("\n")
    for i, l in enumerate(line):
        cv.putText(
            m,
            l,
            pos.astype("int32") + [0, i * lineheight],
            cv.FONT_HERSHEY_SIMPLEX,
            1,
            textcolor,
        )
    return m


class HotkeyManager:
    """
    to piorer ctrl+c than c
    responde no c after doing ctrl+c
    issue:
    1. ctrl+a,ctrl+a and ctrl+a,ctrl+s
        fastly press ctrl+a, and ctrl+s, could cause slightly overlap whose key state is [ctrl,a,s] physically
        that is, [ctrl,a]->[ctrl,a,s]->[ctrl,s]
        so at the second state, both [ctrl+a,ctrl+a] and [ctrl+a,ctrl+s] are triggered
        not severe problem
    """

    @dataclasses.dataclass
    class ContiniousCallHandler:
        prevState: typing.List[bool] = dataclasses.field(default_factory=list)
        countiousPressTime: int = 0
        startRepeatPeriod: int = 10
        useControlOnContiniousPress: bool = True

        @staticmethod
        def __dictEq(a: typing.Dict[int, bool], b: typing.Dict[int, bool]):
            if len(a) != len(b):
                return False
            for k in a.keys():
                if k not in b.keys():
                    return False
                if a[k] != b[k]:
                    return False
            return True

        def updateState(self, newState):
            if HotkeyManager.ContiniousCallHandler.__dictEq(self.prevState, newState):
                self.countiousPressTime += 1
                return self.countiousPressTime >= self.startRepeatPeriod
            else:
                self.countiousPressTime = 0
                self.prevState = newState
                return True

    class hotkeytask:
        key: List[List[int]]

        def __init__(
            self,
            key: int | Iterable,
            foo: Callable[[], None],
            continiousPress: bool = False,
        ) -> None:
            if not isinstance(key, Iterable):
                self.key = [[key]]
            elif not isinstance(key[0], Iterable):
                self.key = [key]
            else:
                self.key = key
            # self.key is like [keyset1=[key1, key2], keyset2=[key3, key4]]
            self.foo = foo
            self.continiousPress = continiousPress
            self.stage = 0

        def tryRespond(self, respond: bool, anyRespondingExceptThis: bool):
            if respond:
                if self.stage == len(self.key) - 1:
                    # progress completed
                    self.stage = 0
                    self.foo()
                else:
                    self.stage += 1
            elif anyRespondingExceptThis:
                self.stage = 0

        def GetNowKey(self):
            return self.key[self.stage]

    @dataclasses.dataclass
    class Key:
        code: int

        def GetKeyDown(self):
            return isKBDown(self.code)

    def __init__(self, hotkeytasklist: List[hotkeytask]):
        keyconcerned = [hka.key for hka in hotkeytasklist]
        keyconcerned = list(itertools.chain.from_iterable(keyconcerned))
        keyconcerned = list(itertools.chain.from_iterable(keyconcerned))
        keyconcerned = deduplicate(keyconcerned)
        self.kc = [HotkeyManager.Key(k) for k in keyconcerned]

        # clear all previous state
        [k.GetKeyDown() for k in self.kc]

        self.hktl = hotkeytasklist
        self.cch = HotkeyManager.ContiniousCallHandler()

    def decideAllHotKey(self) -> List[bool]:
        keystate = {k.code: k.GetKeyDown() for k in self.kc}
        cchblocked = not self.cch.updateState(keystate)

        class respondstate(enum.Enum):
            false = 0
            true = 1
            unknown = 2

        respondtable = [
            respondstate.unknown
            if not cchblocked or hk.continiousPress
            else respondstate.false
            for hk in self.hktl
        ]

        def piorered(a: HotkeyManager.hotkeytask, b: HotkeyManager.hotkeytask):
            def include(a: HotkeyManager.hotkeytask, b: HotkeyManager.hotkeytask):
                for k in b.GetNowKey():
                    if k not in a.GetNowKey():
                        return False
                return True

            # a>b and b<a, not equal
            return include(a, b) and not include(b, a)

        # costly
        piorinfo = [
            [
                aidx
                for aidx, a in enumerate(self.hktl)
                if aidx != bidx and piorered(a, b)
            ]
            for bidx, b in enumerate(self.hktl)
        ]

        def decideRespondState(i):
            # checked
            if respondtable[i] != respondstate.unknown:
                return

            # all key pressed
            if all([keystate[k] for k in self.hktl[i].GetNowKey()]):
                # didnt check piored, check it
                [
                    decideRespondState(p)
                    for p in piorinfo[i]
                    if respondtable[p] == respondstate.unknown
                ]

                # no piored responded
                if all([respondtable[p] == respondstate.false for p in piorinfo[i]]):
                    respondtable[i] = respondstate.true
                else:
                    respondtable[i] = respondstate.false
            else:
                # not respond this
                respondtable[i] = respondstate.false

        for hkidx, hk in enumerate(self.hktl):
            decideRespondState(hkidx)

        assert all([rt != respondstate.unknown for rt in respondtable])
        r = [
            respondtable[hkidx] == respondstate.true
            for hkidx, hk in enumerate(self.hktl)
        ]
        return [
            respondtable[hkidx] == respondstate.true
            for hkidx, hk in enumerate(self.hktl)
        ]

    def doAllDecidedKey(self, decideresult, throwonerr=False, printonerr=False):
        AnyRespondingExceptThis = [
            any([(rb if i != j else False) for j, rb in enumerate(decideresult)])
            for i, ra in enumerate(decideresult)
        ]
        for i in range(len(decideresult)):
            try:
                self.hktl[i].tryRespond(decideresult[i], AnyRespondingExceptThis[i])
            except Exception as e:
                if printonerr:
                    traceback.print_exc()
                if throwonerr:
                    raise e


class DataCollector:
    randNameLen = 10

    def __init__(self, outputpath) -> None:
        self.outputpath = outputpath

    @staticmethod
    def geneName():
        charSet4RandomString = "1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        return randomString(charSet4RandomString, DataCollector.randNameLen)

    def save(self, m, name=None):
        if name is None:
            name = DataCollector.geneName()
        savemat(m, f"{name}", path=self.outputpath)


def AllFileIn(path, includeFileInSubDir=True):
    ret = []
    for dirpath, dir, file in os.walk(path):
        if not includeFileInSubDir and dirpath != path:
            continue
        ret.extend([os.path.join(dirpath, f) for f in file])
    return ret


class StoppableThread:
    """
    derivate from it and override foo()
    """

    def __init__(self, pool: ThreadPoolExecutor = None) -> None:
        self.running: bool = False
        self.pool: ThreadPoolExecutor = (
            pool if pool is not None else ThreadPoolExecutor()
        )
        self.submit = None

    def foo(self) -> None:
        raise NotImplementedError("should never run without overriding foo")

    def getRunning(self) -> bool:
        return self.running

    def go(self) -> None:
        if self.submit is not None:
            return
        self.running = True

        def call() -> None:
            """
            wrapper so can call the passed "self"'s foo
            if not, can never know which overwritten foo should be called
            """
            self.foo()

        self.submit = self.pool.submit(call)

    def stop(self) -> None:
        if self.submit is None:
            return
        self.running = False
        self.submit.result()
        self.submit = None


def ReadTextFile(path):
    with open(path, "r") as f:
        return f.read()


def WriteTextFile(path, text):
    with open(path, "w") as f:
        f.write(text)


def FunctionalWrapper(f: Callable[..., Any]) -> Callable[..., Any]:
    """
    A decorator that wraps a function and returns a new function that calls the original function
    and returns the instance it was called on.

    Args:
        f: The function to wrap.

    Returns:
        The wrapped function.
    """

    def f2(self, *args: Any, **kwargs: Any) -> Any:
        f(self, *args, **kwargs)
        return self

    return f2


class Pipe:
    value: Any = None

    def __init__(self, initValue: Any = None, printStep: bool = False) -> None:
        """
        Initializes a Pipe object.

        Args:
            initValue: The initial value of the Pipe.
            printStep: Whether to print the value after setting it.
        """
        self.printStep = printStep
        self.set(initValue)

    def get(self) -> Any:
        """
        Returns the value of the Pipe.

        Returns:
            The value of the Pipe.
        """
        return self.value

    @FunctionalWrapper
    def set(self, val: Any) -> None:
        """
        Sets the value of the Pipe.

        Args:
            val: The value to set.
        """
        self.value = val
        if self.printStep:
            print(self.value)

    @FunctionalWrapper
    def do(self, foo: Callable[[Any], Any]) -> "Pipe":
        """
        Applies a function to the value of the Pipe and sets the result as the new value.

        Args:
            foo: The function to apply.

        Returns:
            The modified Pipe object.
        """
        self.set(foo(self.get()))
        return self

    def __repr__(self) -> str:
        """
        Returns the string representation of the Pipe.

        Returns:
            The string representation of the Pipe.
        """
        return self.get().__repr__()


def GetTimeString():
    return time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())


def UnfinishedWrapper(msg=None) -> Callable[..., Any]:
    if callable(msg):
        # calling without parens, works both on a class and a function
        foo = msg
        return UnfinishedWrapper()(foo)

    default_msg = "Unfinished"
    if msg is None:
        msg = default_msg

    def f2(foo):
        def f3(*args: Any, **kwargs: Any) -> Any:
            raise NotImplementedError(msg)

        return f3

    return f2


class expparser:
    """
    TODO:
    numlike
        tensor support
        string support
            almost done
            considering impl operator on this type
    add type conversion func
    """

    class TokenType(enum.Enum):
        NUMLIKE = 1
        OPR = 2
        BRA = 3
        KET = 4
        EOF = 5
        IDR = 6

    @dataclasses.dataclass(repr=True)
    class Token:
        type: "expparser.TokenType"
        value: typing.Any
        start: int
        end: int

    class __State(enum.Enum):
        START = 1
        NUM = 2
        NEG = 3
        OPR = 4
        END = 5
        IDR = 6

    @dataclasses.dataclass
    class __OprPriorityLeap:
        pos: int
        pribefore: int
        priafter: int

    class NumLikeUnionUtil:
        class NumLikeException(BaseException):
            pass

        class NumLikeType(enum.Enum):
            NUM = 0
            STR = 1
            LIST = 2
            BOOL = 3

        @staticmethod
        def TypeOf(nl):
            if isinstance(nl, str):
                return expparser.NumLikeUnionUtil.NumLikeType.STR
            elif isinstance(nl, typing.Iterable):
                return expparser.NumLikeUnionUtil.NumLikeType.LIST
            elif isinstance(nl, float):
                return expparser.NumLikeUnionUtil.NumLikeType.NUM
            elif isinstance(nl, bool):
                return expparser.NumLikeUnionUtil.NumLikeType.BOOL
            else:
                raise expparser.NumLikeUnionUtil.NumLikeException()

        # imexplicit conversion
        @staticmethod
        def ToNum(nl):
            t = expparser.NumLikeUnionUtil.TypeOf(nl)
            if t == expparser.NumLikeUnionUtil.NumLikeType.NUM:
                return nl
            elif t == expparser.NumLikeUnionUtil.NumLikeType.BOOL:
                return 1.0 if nl else 0.0
            else:
                raise expparser.NumLikeUnionUtil.NumLikeException()

        @staticmethod
        def ToList(nl):
            t = expparser.NumLikeUnionUtil.TypeOf(nl)
            if t == expparser.NumLikeUnionUtil.NumLikeType.LIST:
                return list(nl)
            elif t == expparser.NumLikeUnionUtil.NumLikeType.STR:
                return [nl]
            elif t == expparser.NumLikeUnionUtil.NumLikeType.NUM:
                return [nl]
            elif t == expparser.NumLikeUnionUtil.NumLikeType.BOOL:
                return [nl]
            else:
                raise expparser.NumLikeUnionUtil.NumLikeException()

        @staticmethod
        def ToBool(nl):
            t = expparser.NumLikeUnionUtil.TypeOf(nl)
            if t == expparser.NumLikeUnionUtil.NumLikeType.NUM:
                return nl > 0
            elif t == expparser.NumLikeUnionUtil.NumLikeType.BOOL:
                return nl
            else:
                raise expparser.NumLikeUnionUtil.NumLikeException()

        @staticmethod
        def ToProperFormFromAny(nl):
            # the adaptor for data to be proper
            # after which there will be no int form, they all stored as float
            # this will deal with more situations, so TypeOf is not proper here
            if isinstance(nl, str):
                return nl
            elif isinstance(nl, typing.Iterable):
                return list(nl)
            elif isinstance(nl, float):
                return nl
            elif isinstance(nl, bool):
                # isinstance(True, int)==True
                return nl
            elif isinstance(nl, int):
                return float(nl)
            else:
                raise expparser.NumLikeUnionUtil.NumLikeException()

    class ParseException(BaseException):
        pass

    BasicConstantLib = {"e": np.e, "pi": np.pi, "true": True, "false": False}

    @staticmethod
    def unpackParaArray(f):
        def f2(a):
            if (
                expparser.NumLikeUnionUtil.TypeOf(a)
                == expparser.NumLikeUnionUtil.NumLikeType.LIST
            ):
                return f(*a)
            else:
                return f(a)

        return f2

    @staticmethod
    def CList(a):
        if (
            expparser.NumLikeUnionUtil.TypeOf(a)
            == expparser.NumLikeUnionUtil.NumLikeType.LIST
        ):
            return a
        else:
            return [a]

    BasicFunctionLib = {
        "sin": unpackParaArray(np.sin),
        "cos": unpackParaArray(np.cos),
        "tan": unpackParaArray(np.tan),
        "asin": unpackParaArray(np.arcsin),
        "acos": unpackParaArray(np.arccos),
        "atan": unpackParaArray(np.arctan),
        "atan2": unpackParaArray(np.arctan),
        "exp": unpackParaArray(np.exp),
        "log": unpackParaArray(np.log),
        "sqrt": unpackParaArray(np.sqrt),
        "abs": unpackParaArray(np.abs),
        "sign": unpackParaArray(np.sign),
        "floor": unpackParaArray(np.floor),
        "ceil": unpackParaArray(np.ceil),
        "round": unpackParaArray(np.round),
        "neg": unpackParaArray(lambda x: -x),
        "iif": unpackParaArray(
            lambda cond, x, y: x if expparser.NumLikeUnionUtil.ToBool(cond) else y
        ),
        "eq": unpackParaArray(lambda x, y, ep=0.001: abs(x - y) < ep),
        "streq": unpackParaArray(lambda x, y: x == y),
        "array": lambda a: a,
        "CStr": unpackParaArray(str),
        "CNum": unpackParaArray(float),
        "CBool": unpackParaArray(bool),
        "CList": unpackParaArray(CList),
        "WrapSingle": lambda a: [a],
    }

    class OprException(BaseException):
        pass

    class _OprType(enum.Enum):
        UNSPECIFIED = 0
        ADD = 1
        SUB = 2
        MUL = 3
        DIV = 4
        POW = 5
        COMMA = 6
        NEG = 7  # available by manual transfering from sub
        NEQ = 8
        EQ = 9
        GT = 10
        GE = 11
        LT = 12
        LE = 13
        NOT = 14
        AND = 15
        OR = 16
        XOR = 17

        @staticmethod
        def __throw_opr_exception(s):
            raise expparser.OprException(f"bad opr {s}")

        def getPriority(self):
            if self == expparser._OprType.COMMA:
                return 1
            elif self in [
                expparser._OprType.OR,
                expparser._OprType.AND,
                expparser._OprType.XOR,
            ]:
                return 2
            elif self in [
                expparser._OprType.GT,
                expparser._OprType.GE,
                expparser._OprType.LT,
                expparser._OprType.LE,
                expparser._OprType.EQ,
                expparser._OprType.NEQ,
            ]:
                return 3
            elif self in [expparser._OprType.ADD, expparser._OprType.SUB]:
                return 4
            elif self in [expparser._OprType.MUL, expparser._OprType.DIV]:
                return 5
            elif self == expparser._OprType.POW:
                return 6
            else:
                expparser._OprType.__throw_opr_exception(self)

        @staticmethod
        def fromStr(s):
            dict = {
                "+": expparser._OprType.ADD,
                "-": expparser._OprType.SUB,
                "*": expparser._OprType.MUL,
                "/": expparser._OprType.DIV,
                "^": expparser._OprType.POW,
                ",": expparser._OprType.COMMA,
                "=": expparser._OprType.EQ,
                "!=": expparser._OprType.NEQ,
                ">": expparser._OprType.GT,
                ">=": expparser._OprType.GE,
                "<": expparser._OprType.LT,
                "<=": expparser._OprType.LE,
                "!": expparser._OprType.NOT,
                "&": expparser._OprType.AND,
                "|": expparser._OprType.OR,
                "^^": expparser._OprType.XOR,
            }
            if s not in dict:
                expparser._OprType.__throw_opr_exception(s)
            return dict[s]

        def do(self, arg: typing.List["expparser.Token"], serial=0):
            # take several tokens and return numlike
            if self == expparser._OprType.COMMA:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                # comma behaves differently with serial==0 and serial==other
                if serial == 0:
                    return [arg[0].value] + [arg[1].value]
                else:
                    return expparser.NumLikeUnionUtil.ToList(arg[0].value) + [
                        arg[1].value
                    ]
            elif self == expparser._OprType.ADD:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) + expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.SUB:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) - expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.MUL:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) * expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.DIV:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) / expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.POW:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) ** expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.NEG:
                assert arg[0].type == expparser.TokenType.NUMLIKE
                return -expparser.NumLikeUnionUtil.ToNum(arg[0].value)
            elif self == expparser._OprType.NEQ:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) != expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.EQ:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) == expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.GT:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) > expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.GE:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) >= expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.LT:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) < expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.LE:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToNum(
                    arg[0].value
                ) <= expparser.NumLikeUnionUtil.ToNum(arg[1].value)
            elif self == expparser._OprType.NOT:
                assert arg[0].type == expparser.TokenType.NUMLIKE
                return not expparser.NumLikeUnionUtil.ToBool(arg[0].value)
            elif self == expparser._OprType.AND:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToBool(
                    arg[0].value
                ) and expparser.NumLikeUnionUtil.ToBool(arg[1].value)
            elif self == expparser._OprType.OR:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToBool(
                    arg[0].value
                ) or expparser.NumLikeUnionUtil.ToBool(arg[1].value)
            elif self == expparser._OprType.XOR:
                assert (
                    arg[0].type == expparser.TokenType.NUMLIKE
                    and arg[1].type == expparser.TokenType.NUMLIKE
                )
                return expparser.NumLikeUnionUtil.ToBool(
                    arg[0].value
                ) ^ expparser.NumLikeUnionUtil.ToBool(arg[1].value)
            else:
                expparser._OprType.__throw_opr_exception(self)

        def isUnary(self):
            return self in [expparser._OprType.NEG, expparser._OprType.NOT]

    @staticmethod
    def __NextToken(s, i=0):
        @dataclasses.dataclass
        class matcher:
            exp: str
            tokenType: expparser.TokenType

            def tryMatch(self, s, i):
                match = re.match(self.exp, s[i:])
                if not match:
                    return None
                end = match.span()[1] + i
                return expparser.Token(self.tokenType, s[i:end], i, end)

        matcherList = [
            matcher(
                exp=r"^(<=)|(>=)|(\^\^)|(!=)", tokenType=expparser.TokenType.OPR
            ),  # two width operator, match before single widthed ones to get priority
            matcher(
                exp=r"^[*/+\-^,=<>&|]", tokenType=expparser.TokenType.OPR
            ),  # single width operator
            matcher(exp=r"^[0-9]+(\.[0-9]+)?", tokenType=expparser.TokenType.NUMLIKE),
            matcher(exp=r'^".*?"', tokenType=expparser.TokenType.NUMLIKE),
            matcher(exp=r"^[A-Za-z_][A-Za-z0-9_]*", tokenType=expparser.TokenType.IDR),
            matcher(exp=r"^\(", tokenType=expparser.TokenType.BRA),
            matcher(exp=r"^\)", tokenType=expparser.TokenType.KET),
            matcher(exp=r"^$", tokenType=expparser.TokenType.EOF),
        ]
        # "the \" string"
        for m in matcherList:
            ret = m.tryMatch(s, i)
            if ret:
                if ret.type == expparser.TokenType.NUMLIKE:
                    # here is only num and str without other numlike types
                    if len(ret.value) >= 2 and ret.value[0] == '"':
                        # its string
                        strstart = ret.start
                        strbuffer = ""
                        while True:
                            if len(ret.value) > 2 and ret.value[-2] == "\\":
                                # cancel the former " and \, add " back
                                strbuffer += ret.value[1:-2] + '"'
                                # move on to the next section
                                ret = m.tryMatch(s, ret.end - 1)
                            else:
                                strbuffer += ret.value[1:-1]
                                break
                        ret.start = strstart
                        ret.value = strbuffer
                    else:
                        ret.value = float(ret.value)
                elif ret.type == expparser.TokenType.OPR:
                    ret.value = expparser._OprType.fromStr(ret.value)
                return ret
        raise expparser.ParseException(f"unexpected token at {i}")

    @dataclasses.dataclass
    class __ExpParserResult:
        val: typing.Any
        end: int

    @staticmethod
    def __expparse_recursive(s, varList: typing.Dict, funcList: typing.Dict, i=0):
        # fsm fields
        state = expparser.__State.START
        token: expparser.Token = None
        # never modify peekToken
        peekToken = expparser.__NextToken(s, i)

        # buffer
        tokenList: typing.List[expparser.Token] = []

        # for operator priority
        oprRisingBeginPosList: typing.List[expparser.__OprPriorityLeap] = [
            expparser.__OprPriorityLeap(0, 0, 0)
        ]

        def RaiseTokenException(token: expparser.Token):
            raise expparser.ParseException(
                f'unexpected {token.type}("{token.value}") at {token.start}'
            )

        def ClearOprSectionAssumingPeer(begin, end):
            nonlocal tokenList, oprRisingBeginPosList
            i = begin
            val = tokenList[i]
            i += 1
            peerPriority = None
            serial = 0
            while True:
                if i >= end:
                    break
                opr = tokenList[i]
                i += 1
                val2 = tokenList[i]
                i += 1
                if peerPriority:
                    assert opr.value.getPriority() == peerPriority
                else:
                    peerPriority = opr.value.getPriority()
                val.value = opr.value.do(serial=serial, arg=[val, val2])
                serial += 1
            # replace the peer section into its result
            tokenList = (
                tokenList[:begin]
                + [expparser.Token(expparser.TokenType.NUMLIKE, val.value, 0, 0)]
                + tokenList[end:]
            )
            RemapToken()
            oprRisingBeginPosList.pop()
            return val

        def DealWithBra():
            nonlocal token, peekToken, state, tokenList
            subresult = expparser.__expparse_recursive(s, varList, funcList, token.end)
            tokenList.pop()  # remove the bra
            tokenList.append(
                expparser.Token(expparser.TokenType.NUMLIKE, subresult.val, 0, 0)
            )
            RemapToken()
            peekToken = expparser.__NextToken(s, subresult.end)
            state = expparser.__State.NUM

        def DealWithIdentifier():
            nonlocal token, peekToken, state, tokenList
            if peekToken.type == expparser.TokenType.BRA:
                # its a call
                fooName = token.value
                assert fooName in funcList
                # manually move on
                MoveForwardToNextToken()
                DealWithBra()
                para = tokenList[-1].value
                tokenList[-1].value = expparser.NumLikeUnionUtil.ToProperFormFromAny(
                    funcList[fooName](para)
                )
                tokenList = tokenList[:-2] + tokenList[-1:]  # remove the func name
                RemapToken()
            else:
                # its a var
                assert token.value in varList
                # overwrite the identifier with num
                token.value = expparser.NumLikeUnionUtil.ToProperFormFromAny(
                    varList[token.value]
                )
                token.type = expparser.TokenType.NUMLIKE
            state = expparser.__State.NUM

        def MoveForwardToNextToken():
            nonlocal token, peekToken, tokenList
            token = peekToken
            peekToken = expparser.__NextToken(s, token.end)
            tokenList.append(token)
            # RemapToken() # not essential here
            # modifying token also applies on tokenList[-1]

        def RemapToken():
            nonlocal tokenList, token
            # reset token to last of token list after modifying tokenList structure
            if len(tokenList) != 0:
                token = tokenList[-1]
            else:
                # calling any member on this will cause exception
                token = None

        # the fsm illustrated in notebook
        while True:
            MoveForwardToNextToken()
            if state == expparser.__State.START or state == expparser.__State.OPR:
                # expecting numlike, but possible to meet bra, identifier, or unary operator
                def DealWithUnaryOperatorIfNeeded():
                    nonlocal state, token, peekToken, tokenList
                    while True:
                        # do in a loop to deal with nested unary
                        if (
                            len(tokenList) >= 2
                            and tokenList[-2].type == expparser.TokenType.OPR
                            and tokenList[-2].value.isUnary()
                        ):
                            # before it is unary operator, do it
                            token.value = tokenList[-2].value.do(arg=[token])
                            tokenList = tokenList[:-2] + [token]
                            RemapToken()
                        else:
                            break

                if token.type == expparser.TokenType.BRA:
                    DealWithBra()
                    DealWithUnaryOperatorIfNeeded()
                elif token.type == expparser.TokenType.NUMLIKE:
                    state = expparser.__State.NUM
                    DealWithUnaryOperatorIfNeeded()
                elif token.type == expparser.TokenType.IDR:
                    DealWithIdentifier()
                    DealWithUnaryOperatorIfNeeded()
                elif token.type == expparser.TokenType.OPR:
                    # maybe unary operator
                    state = expparser.__State.OPR
                    if token.value == expparser._OprType.SUB:
                        # to neg
                        tokenList[-1].value = expparser._OprType.NEG
                    if not token.value.isUnary():
                        RaiseTokenException(token)
                    # will be calced after read the num
                else:
                    RaiseTokenException(token)
            elif state == expparser.__State.NUM:
                if token.type == expparser.TokenType.OPR:
                    state = expparser.__State.OPR
                    lastOprPrior = oprRisingBeginPosList[-1].priafter
                    opr = token.value.getPriority()
                    if opr > lastOprPrior:
                        oprRisingBeginPosList.append(
                            expparser.__OprPriorityLeap(
                                len(tokenList) - 2, lastOprPrior, opr
                            )
                        )  # -2 for num and opr
                    elif opr < lastOprPrior:
                        while True:
                            if opr >= oprRisingBeginPosList[-1].priafter:
                                break
                            # clear since last rising, until flat
                            # len(tokenList)-1 since the lowering opr has been appended
                            ClearOprSectionAssumingPeer(
                                oprRisingBeginPosList[-1].pos, len(tokenList) - 1
                            )
                        if opr > oprRisingBeginPosList[-1].priafter:
                            # new opr is the new rising
                            oprRisingBeginPosList.append(
                                expparser.__OprPriorityLeap(
                                    len(tokenList) - 2,
                                    oprRisingBeginPosList[-1].priafter,
                                    opr,
                                )
                            )
                elif (
                    token.type == expparser.TokenType.KET
                    or token.type == expparser.TokenType.EOF
                ):
                    state = expparser.__State.END
                    expendpos = token.end
                    tokenList.pop()  # remove the eof or ket
                    RemapToken()
                    # clear all
                    while True:
                        if len(oprRisingBeginPosList) == 0:
                            break
                        ClearOprSectionAssumingPeer(
                            oprRisingBeginPosList[-1].pos, len(tokenList)
                        )
                    return expparser.__ExpParserResult(tokenList[0].value, expendpos)
                else:
                    RaiseTokenException(token)
            elif state == expparser.__State.END:
                RaiseTokenException(token)

    @staticmethod
    def parseelement(s):
        i = 0
        tokenList: typing.List[expparser.Token] = []
        while True:
            token = expparser.__NextToken(s, i)
            tokenList.append(token)
            i = token.end
            if token.type == expparser.TokenType.EOF:
                break
        var = []
        func = []
        for i, tk in enumerate(tokenList):
            if tk.type == expparser.TokenType.IDR:
                # peek
                # wont index overflow cuz there is eof and eof is not identifier
                if tokenList[i + 1].type == expparser.TokenType.BRA:
                    func.append(tk.value)
                else:
                    var.append(tk.value)
        return var, func

    @staticmethod
    def expparse(s, var={}, func={}):
        return expparser.__expparse_recursive(s, var, func).val

    @staticmethod
    def test():
        var = {**expparser.BasicConstantLib}
        func = {**expparser.BasicFunctionLib}

        @dataclasses.dataclass
        class TestUnit:
            expression: str
            expected: str = "unspecified"
            result: str = ""

        exp: typing.List[TestUnit] = [
            TestUnit(r"sin(pi/2)", "1.0"),
            TestUnit(r"1+2^2*2+--1", "10.0"),
            TestUnit(r'streq("test \" str","test \" str")', "True"),
            TestUnit(r"eq(1+0.1,1)", "False"),
            TestUnit(r"eq(1+0.1,1,0.2)", "True"),
            TestUnit(r"1!=2^^2>=3", "True"),
            TestUnit(r"CList(1))", "[1.0]"),
            TestUnit(r"CBool(1))", "True"),
            TestUnit(r"CBool(0))", "False"),
            TestUnit(r'streq(CStr(1),"1.0")', "True"),
            TestUnit(r"CStr(true)", "True"),
            TestUnit(r'CNum("1.23")+1', "2.23"),
            TestUnit(r"CNum(true)+1", "2.0"),
            TestUnit(r"array(array(1,0),array(0,1))", "[[1.0, 0.0], [0.0, 1.0]]"),
            TestUnit(r"WrapSingle(WrapSingle(1))", "[[1.0]]"),
        ]
        unpassed: typing.List[TestUnit] = []

        def splitline():
            print("#" * 30)

        for e in exp:
            result = expparser.expparse(
                e.expression,
                var=var,
                func=func,
            )
            if str(result) == e.expected:
                ifpass = True
            else:
                ifpass = False
                e.result = result
                unpassed.append(e)
            print(
                f"""
exp: {e.expression}
    elements: {expparser.parseelement(e.expression)}
    expected: {e.expected}
    result: {result}
    {"pass" if ifpass else "fail"}
"""[
                    1:-1
                ]
            )
            splitline()
        if len(unpassed) == 0:
            print("all passed")
        else:
            print("unpassed")
            splitline()
            for u in unpassed:
                print(
                    f"""
{u.expression}
    expected: {u.expected}
    result: {u.result}
"""[
                        1:-1
                    ]
                )
                splitline()


def sleepuntil(con: Callable, dt=0.01):
    while not con():
        time.sleep(dt)


class perf_statistic:
    def __init__(self, startnow=False):
        self.clear()
        if startnow:
            self.start()

    def clear(self):
        self._starttime = None
        self._stagedtime = 0
        self._cycle = 0

    def start(self):
        self._starttime = time.perf_counter()

    def countcycle(self):
        self._cycle += 1

    def stop(self):
        if self._starttime is None:
            return
        self._stagedtime += self._timeCurrentlyCounting()
        self._starttime = None

    def aveTime(self):
        return self.time() / self._cycle if self._cycle > 0 else 0

    def _timeCurrentlyCounting(self):
        return (
            time.perf_counter() - self._starttime if self._starttime is not None else 0
        )

    def time(self):
        return self._stagedtime + self._timeCurrentlyCounting()


class fpsmanager:
    def __init__(self, fps=60):
        self.lt = time.perf_counter()
        self.frametime = 1 / fps

    def WaitUntilNextFrame(self):
        sleepuntil(
            lambda: time.perf_counter() - self.lt > self.frametime,
            dt=0.5 * self.frametime,
        )
        self.SetToNextFrame()

    def CheckIfTimeToDoNextFrame(self) -> bool:
        """
        usage
        if fpsmanager.CheckIfTimeToDoNextFrame():
            fpsmanager.SetToNextFrame()
            do your task here
        used in doing stuff peroidically, but in another loop with different peroid
        so have to check if it is time to do it
        """
        result = time.perf_counter() - self.lt > self.frametime
        return result

    def SetToNextFrame(self):
        self.lt = time.perf_counter()


"""
windows
"""

from ctypes import windll, byref, c_ubyte
from ctypes.wintypes import RECT, HWND
import win32api
import win32con
import win32gui
import win32ui


def getWTHwnd():
    ret = win32gui.FindWindow("DagorWClass", None)
    if ret == win32con.NULL:
        raise Exception("FindWindow() failed")
    return ret


class screenshoter:
    # reconstructed with no mfc

    def __init__(self, hwnd=0):
        self.wthwnd = hwnd
        # r = RECT()
        # windll.user32.GetClientRect(self.hwnd, byref(r))
        # self.res=[r.right,r.bottom]
        self.res = [1920, 1080]
        self.wtdc = windll.user32.GetDC(self.wthwnd)
        self.mydc = windll.gdi32.CreateCompatibleDC(self.wtdc)
        self.mybitmap = windll.gdi32.CreateCompatibleBitmap(
            self.wtdc, self.res[0], self.res[1]
        )
        windll.gdi32.SelectObject(self.mydc, self.mybitmap)

    def __del__(self):
        windll.gdi32.DeleteObject(self.mybitmap)
        windll.gdi32.DeleteObject(self.mydc)
        windll.user32.ReleaseDC(self.wthwnd, self.wtdc)

    def shot(self):
        if (
            windll.gdi32.BitBlt(
                self.mydc,
                0,
                0,
                self.res[0],
                self.res[1],
                self.wtdc,
                0,
                0,
                win32con.SRCCOPY,
            )
            == 0
        ):
            raise BaseException("bad shot, {}".format(windll.kernel32.GetLastError()))
        # 截图是BGRA排列，因此总元素个数需要乘以4
        total_bytes = self.res[0] * self.res[1] * 4
        buffer = bytearray(total_bytes)
        byte_array = c_ubyte * total_bytes
        windll.gdi32.GetBitmapBits(
            self.mybitmap, total_bytes, byte_array.from_buffer(buffer)
        )
        return np.frombuffer(buffer, dtype=np.uint8).reshape(
            self.res[1], self.res[0], 4
        )

    def shotbgr(self):
        return self.shot()[:, :, :3]

    def shotgray(self):
        return cv.cvtColor(self.shot(), cv.COLOR_BGRA2GRAY)


def activeWindow(hwnd):  # 窗口置顶
    win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
    win32gui.SetWindowPos(
        hwnd,
        win32con.HWND_TOP,
        0,
        0,
        0,
        0,
        win32con.SWP_NOSIZE | win32con.SWP_SHOWWINDOW,
    )
    win32gui.SetForegroundWindow(hwnd)


def bootAsAdmin(file):
    # pass in __file__
    quotedpy = '"' + file + '"'
    ret = ctypes.windll.shell32.ShellExecuteW(
        None, "runas", sys.executable, quotedpy, None, 1
    )


def setadmin(file):
    def is_admin():
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    if not is_admin():
        bootAsAdmin(file)
        exit()


class fullScrHUD:
    """
    #init
    hud=fullScrHUD()
    hud.setup()
    ...
    #main loop
    #preparing to show
    hud.clear
    #init m1
    m1=hud.getblankscreen()
    #draw sth on m1
    #then add m1 to hud
    hud.addcontent(m1)
    #or overwrite directly
    hud.writecontent(m1)
    #do the same for m2...
    ...
    #ask hud to show
    hud.update()
    ...
    #on exit
    hud.stop()
    """

    def __init__(self) -> None:
        self.resolution = [1080, 1920]
        self.terminate = False
        self.hwnd = 0
        self.m2show = self.getblankscreenwithalfa()
        self.m2draw = self.getblankscreenwithalfa()
        # set m2draw

    def setup(self):
        def WndProc(hwnd, msg, wParam, lParam):
            if msg == win32con.WM_PAINT:
                rect = win32gui.GetClientRect(hwnd)
                hdc, ps = win32gui.BeginPaint(hwnd)
                # background
                # hbr=win32gui.CreateSolidBrush(0x0000000)
                # win32gui.FillRect(hdc,rect,hbr)
                # win32gui.DeleteObject(hbr)

                w = self.resolution[1]
                h = self.resolution[0]
                mfcDC = win32ui.CreateDCFromHandle(hdc)
                hcdc = mfcDC.CreateCompatibleDC()
                BitMap = win32ui.CreateBitmap()
                BitMap.CreateCompatibleBitmap(mfcDC, w, h)
                ctypes.WinDLL("gdi32.dll").SetBitmapBits(
                    BitMap.GetHandle(), w * h * 4, self.m2show.tobytes()
                )
                hcdc.SelectObject(BitMap)
                # myblendfunc=(win32con.AC_SRC_OVER,0,255,win32con.AC_SRC_ALPHA)
                # win32gui.AlphaBlend(hdc,0,0,w, h , hcdc.GetHandleAttrib(), 0,0,w, h,myblendfunc)
                win32gui.BitBlt(
                    hdc, 0, 0, w, h, hcdc.GetHandleAttrib(), 0, 0, win32con.SRCCOPY
                )
                win32gui.DeleteObject(BitMap.GetHandle())
                hcdc.DeleteDC()
                win32gui.EndPaint(hwnd, ps)
            elif msg == win32con.WM_SIZE:
                win32gui.InvalidateRect(hwnd, None, True)
            elif msg == win32con.WM_DESTROY:
                win32gui.PostQuitMessage(0)
            if self.terminate:
                win32gui.PostQuitMessage(0)
            return win32gui.DefWindowProc(hwnd, msg, wParam, lParam)

        def mainloop():
            wc = win32gui.WNDCLASS()
            hbrush = win32gui.CreateSolidBrush(0)
            wc.hbrBackground = hbrush
            wc.lpszClassName = "Python on Windows"
            wc.lpfnWndProc = WndProc
            reg = win32gui.RegisterClass(wc)

            hwnd = win32gui.CreateWindow(
                reg,
                "Python",
                win32con.WS_POPUP,
                0,
                0,
                self.resolution[1],
                self.resolution[0],
                0,
                0,
                0,
                None,
            )
            win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
            win32gui.UpdateWindow(hwnd)

            win32gui.SetWindowLong(
                hwnd,
                win32con.GWL_EXSTYLE,
                win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)
                + win32con.WS_EX_LAYERED
                + win32con.WS_EX_NOACTIVATE,
            )
            win32gui.SetLayeredWindowAttributes(hwnd, 0, 0, win32con.LWA_COLORKEY)
            win32gui.SetWindowPos(
                hwnd,
                win32con.HWND_TOPMOST,
                0,
                0,
                0,
                0,
                win32con.SWP_SHOWWINDOW + win32con.SWP_NOSIZE + win32con.SWP_NOMOVE,
            )
            windll.user32.SetWindowDisplayAffinity(
                hwnd, 0x11
            )  # WDA_EXCLUDEDFROMCAPUTURE
            # can not use WDA_MONITOR, or get totally black screen

            self.hwnd = hwnd
            win32gui.PumpMessages()

        t1 = threading.Thread(target=mainloop, args=())
        t1.start()
        time.sleep(1)

    def writecontent(self, lt, content):
        self.m2draw[
            lt[0] : lt[0] + content.shape[0],
            lt[1] : lt[1] + content.shape[1],
            : content.shape[2],
        ] = content

    def clear(self):
        self.m2draw[:, :, :] = 0

    def getblankscreen(self):
        return np.zeros(
            self.resolution
            + [
                3,
            ],
            np.uint8,
        )

    def getblankscreenwithalfa(self):
        return np.zeros(
            self.resolution
            + [
                4,
            ],
            np.uint8,
        )

    def addcontentwithalfa(self, m):
        self.m2draw += m

    def addcontent(self, m):
        alp = 255 * np.ones_like(m[:, :, 0:1])
        self.m2draw += np.concatenate((m, alp), 2)

    def update(self):
        # swap
        tmp = self.m2show
        self.m2show = self.m2draw
        self.m2draw = tmp

        # needless to clear cuz entire screen will be set
        win32gui.InvalidateRect(self.hwnd, None, False)

    def stop(self):
        self.terminate = True
        self.update()

    @staticmethod
    def reverseMat2FitWindowsCoor(m):
        return np.flip(m, axis=0)


def isKBDown(k):
    return bool(win32api.GetAsyncKeyState(k) and 0x1)


def isKBDownNow(k):
    return win32api.GetAsyncKeyState(k) and 0x8000


"""
numpy
"""

import numpy as np

np.seterr(all="raise")


def forceviewmaxmin(m):
    ma = m.max()
    mi = m.min()
    print("ma", ma)
    print("mi", mi)
    pass


def arrayshift(a, n, fill=np.nan):
    # n positive as right
    if n == 0:
        return a
    elif n > 0:
        # a[:-n] will not work as intended on n==0
        return np.concatenate((np.full(n, fill), a[:-n]))
    else:
        return np.concatenate((a[-n:], np.full(-n, fill)))


def integral(dx, x0, keepXM1=False):
    # keepXM1 to keep the last element in x, or deprecate it, cuz its nan if generated from derivative
    x = np.array(list(itertools.accumulate(dx, lambda t, e: t + e)))
    x = np.concatenate([[x0], x if keepXM1 else x[:-1]])
    return x


def derivative(x):
    return arrayshift(x, -1) - x


def summonCard(inteprob, generator=None):
    # summon from card pool
    # impl using np
    # norm
    prob = np.array(inteprob, dtype=np.float32)
    prob /= prob.sum()
    if generator is None:
        return np.random.choice(np.arange(len(prob)), p=prob)
    else:
        return generator.choice(np.arange(len(prob)), p=prob)


class ZFunc:
    """
    x1x2 at any order
    """

    def __init__(self, x1, y1, x2, y2) -> None:
        if x1 < x2:
            # [lower on x or higher on x, x or y]
            self.pt = np.array([[x1, y1], [x2, y2]])
        else:
            self.pt = np.array([[x2, y2], [x1, y1]])
        self.slope = (self.pt[1, 1] - self.pt[0, 1]) / (
            self.pt[1, 0] - self.pt[0, 0] + 0.0001
        )
        self.bias = self.pt[0, 1] - self.pt[0, 0] * self.slope

    def __CallOnNDArray(self, x: np.ndarray):
        y = self.slope * x + self.bias
        y[x < self.pt[0, 0]] = self.pt[0, 1]
        y[x > self.pt[1, 0]] = self.pt[1, 1]
        return y

    def __CallOnNum(self, x):
        if x < self.pt[0, 0]:
            y = self.pt[0, 1]
        elif x > self.pt[1, 0]:
            y = self.pt[1, 1]
        else:
            y = self.slope * x + self.bias
        return y

    def __call__(self, x):
        if type(x) is np.ndarray:
            return self.__CallOnNDArray(x)
        else:
            return self.__CallOnNum(x)


def randomString(charset, length):
    return "".join(
        [
            charset[i]
            for i in np.random.choice(range(len(charset)), length, replace=True)
        ]
    )


"""
opencv
"""

import cv2 as cv


def savemat(m, name=None, path=None, autorename=True):
    if path is None:
        path = r"./output/"
    defaultName = "unnamed"
    defaultSuffix = ".png"
    if name is None:
        name = defaultName + defaultSuffix
    namesplit = os.path.splitext(name)
    name, suffix = str(namesplit[0]), str(namesplit[1])
    if len(suffix) == 0:
        suffix = defaultSuffix

    if not os.path.exists(path):
        os.makedirs(path)
    totalpath = os.path.join(path, name + suffix)
    # find suitable name
    if autorename and os.path.exists(totalpath):
        suffix_idx = 0
        while True:
            suffix_idx += 1
            newname = "{}-{}".format(name, suffix_idx)
            totalpath = os.path.join(path, newname + suffix)
            if not os.path.exists(totalpath):
                break

    if not cv.imwrite(totalpath, m):
        raise IOError(f"Bad write {totalpath}")


def savematn(m: np.ndarray, name=None, path=None):
    mtmp = m.copy()
    cv.normalize(mtmp, mtmp, 0, 255, cv.NORM_MINMAX)
    savemat(mtmp, name, path)


def savematflt(m, multiplier=255, name=None, path=None):
    savemat(multiplier * m, name, path)


def regionsum(m, size, mask=None):
    if m.size <= 0:
        return m
    if mask is not None:
        mask[mask > 0] = 1
    if len(m.shape) > 2 and mask is not None:  # with channel dim
        mask = mask.reshape(mask.shape + (1,))
    return cv.filter2D(m if mask is None else m * mask, -1, np.ones(size, np.float32))


def regionave(m, size, mask=None, notConsiderMaskInDenominator=True):
    """
    if notConsiderMaskInDenominator:
    denominator will not consider mask and boundary and be size[0]*size[1]
    else:
    denominator will be #pix nearby on mask
    u may ask mask==None does the same as notConsiderMaskInDenominator==True
    but if u want to use mask and dont want to be constrained by boundary. unimplemented though
    """

    if m.size <= 0:
        return m
    if mask is not None:
        mask = np.copy(mask)
        mask[mask > 0] = 1
    if mask is None or notConsiderMaskInDenominator:
        denominator = size[0] * size[1]
    else:
        denominator = regionsum(mask, size) + 0.01
        if len(m.shape) > 2:  # m with channel dim
            denominator = denominator.reshape(denominator.shape + (1,))
    return regionsum(m, size, mask) / denominator


def density(p, size):
    return regionave(p.astype("float"), size)


def densityfilter(p, size, thresh):
    dence = density(p, size)
    return np.logical_and(p, dence >= thresh)


rgb2hsvmat = np.array(
    [
        [
            [np.cos(0), np.cos(2 / 3 * np.pi), np.cos(4 / 3 * np.pi)],
            [np.sin(0), np.sin(2 / 3 * np.pi), np.sin(4 / 3 * np.pi)],
            [1, 0, 0],
        ],
        [
            [np.cos(0), np.cos(2 / 3 * np.pi), np.cos(4 / 3 * np.pi)],
            [np.sin(0), np.sin(2 / 3 * np.pi), np.sin(4 / 3 * np.pi)],
            [0, 1, 0],
        ],
        [
            [np.cos(0), np.cos(2 / 3 * np.pi), np.cos(4 / 3 * np.pi)],
            [np.sin(0), np.sin(2 / 3 * np.pi), np.sin(4 / 3 * np.pi)],
            [0, 0, 1],
        ],
    ]
)
hsv2rgbmat = [np.linalg.inv(m) for m in rgb2hsvmat]


def hsv2rgb(hsv):
    h, s, v = hsv
    h = h * np.pi / 180
    xyv = np.array([s * np.cos(h), s * np.sin(h), v])

    # find the corresponding case
    for c, m in enumerate(hsv2rgbmat):
        rgb = m @ xyv
        if np.argmax(rgb) == c:
            return rgb

    # not possible, theoretically
    return np.array((0, 0, 0))

    # to view all solutions
    # rgbs=np.zeros([3,3])
    # for c,m in enumerate(mats):
    #     rgbs[c]=np.linalg.inv(m)@xyv
    # return rgbs


def rgb2hsv(rgb):
    xyv = rgb2hsvmat[np.argmax(rgb)] @ rgb
    x, y, v = xyv
    hsv = np.array([180 / np.pi * np.arctan2(y, x), np.sqrt(x**2 + y**2), v])
    return hsv


def rgb2bgr(rgb):
    m = np.array([[0, 0, 1], [0, 1, 0], [1, 0, 0]])
    return m @ rgb


def convolve_norm(m, k):
    summer = np.ones_like(k)
    mag = np.sqrt(cv.filter2D(m**2, -1, summer))
    ret = cv.filter2D(m, -1, k)
    return ret / mag


def hsv2opencv8bithsv(hsv):
    return np.array([0.5, 2.55, 2.55]) * np.array(hsv)


def outputlines2mat2(m, pos, content, textcolor=[255, 255, 255], lineinterval=10):
    # different impl., ret with content bounding box
    pos = np.array(pos).astype("int")
    line = content.split("\n")
    yoffset = 0
    xmax = 0
    fontFace = cv.FONT_HERSHEY_DUPLEX
    fontScale = 1
    thickness = 1
    for i, l in enumerate(line):
        size = np.array(cv.getTextSize(l, fontFace, fontScale, thickness)[0])
        yoffset += size[1] + lineinterval if i != 0 else size[1]
        if xmax < size[0]:
            xmax = size[0]
        m = cv.putText(
            m,
            l,
            pos + [0, yoffset],
            fontFace,
            fontScale,
            textcolor,
            thickness=thickness,
        )
    box = [pos, pos + [xmax, yoffset]]
    return m, box


def aPicWithText(
    content, maxsize=[1080, 1920], textcolor=[255, 255, 255], lineinterval=10
):
    m = np.zeros(
        maxsize
        + [
            3,
        ],
        np.uint8,
    )
    m, bbox = outputlines2mat2(m, np.array([0, 0]), content, textcolor, lineinterval)
    mshape = np.array(bbox[1]) + [0, 8]  # ret wrong for unknown reason
    m = m[: mshape[1], : mshape[0]]
    m = addShadow2HUD(m)
    return m


def addShadow2HUD(m, thickness=2, color=50):
    gray = cv.cvtColor(m, cv.COLOR_BGR2GRAY)
    kernelshape = 2 * thickness + 1
    edgekernel = np.ones([kernelshape, kernelshape])
    edgekernel[thickness, thickness] = -100  # anchor pix must be black
    # edgekernel=np.array([
    #     [1,1,1],
    #     [1,-80,1],
    #     [1,1,1],
    # ])
    edge = cv.filter2D(gray, -1, edgekernel)
    edge = cv.threshold(edge, 0, 1, cv.THRESH_BINARY)[1]
    edge = edge.reshape(edge.shape + (1,))
    return m + edge * color


def getDemonstrationImg():
    x = np.linspace(0, 5 * 2 * np.pi, 100, dtype=np.float32).reshape(1, -1)
    y = x.T
    demo = np.sin(x + y)
    demo = ZFunc(0, 0.25, 0, 0.75)(demo) * 255
    return demo


"""
xls
"""

import openpyxl as opx


def save_list_to_xls(data_list, filename, sheetname=None):
    # Create a new workbook
    wb = opx.Workbook()

    # Select the active worksheet
    if sheetname is None:
        ws = wb.active
    else:
        ws = wb.create_sheet(sheetname)

    # Iterate over the list and write each item to a new row
    for row, item in enumerate(data_list):
        ws.cell(row=row + 1, column=1, value=item)

    # Save the workbook to the specified filename
    wb.save(filename)


def Xls2ListList(path=None, sheetname=None, killNones=True):
    if path is None:
        path = r"eles.in.xlsx"
    xls = opx.load_workbook(path)
    if sheetname is None:
        sheet = xls.active
    else:
        sheet = xls[sheetname]

    ret = [[ele.value for ele in ln] for ln in (sheet.rows)]
    if killNones:
        ret = [l for l in ret if any([e is not None for e in l])]
    return ret
